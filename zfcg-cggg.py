import requests
import pandas as pd
from datetime import datetime, timedelta
import time
import json
import re
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
# 程序优化为支持动态日期范围（默认抓取近30天的数据）江西公共资源交易中心政府采购类的“招标公告”数据。用“医院”等关键字筛选后的医疗相关的
# 采购数据，含设备招标。
# 配置参数
URL = "https://www.jxsggzy.cn/XZinterface/rest/esinteligentsearch/getFullTextDataNew"
HEADERS = {
    "Content-Type": "application/json;charset=UTF-8",
    "User-Agent": "Mozilla/5.0",
    "Origin": "https://www.jxsggzy.cn",
    "Referer": "https://www.jxsggzy.cn/",
    "Accept": "application/json, text/plain, */*"
}

KEYWORDS = ["医院", "卫生院", "卫生健康", "疾控", "医疗", "妇幼", "总院", "体检"]


def generate_payload(start_date, end_date, page=0, page_size=10):
    return {
        "token": "",
        "pn": page,
        "rn": page_size,
        "sdt": "",
        "edt": "",
        "wd": "",
        "inc_wd": "",
        "exc_wd": "",
        "fields": "",
        "cnum": "",
        "sort": json.dumps({"webdate": "desc"}),
        "ssort": "",
        "cl": 500,
        "terminal": "",
        "condition": [{
            "fieldName": "categorynum",
            "equal": "002006001",
            "isLike": True,
            "likeType": 2
        }],
        "time": [{
            "fieldName": "webdate",
            "startTime": start_date + " 00:00:00",
            "endTime": end_date + " 00:00:00"
        }],
        "highlights": "",
        "statistics": None,
        "unionCondition": [],
        "accuracy": "",
        "noParticiple": "1",
        "searchRange": None,
        "noWd": True
    }


def extract_detail_fields(content):
    bid_time = re.search(r"并于\s*(\d{4}年\d{1,2}月\d{1,2}日\s*\d{1,2}点\d{1,2}分)", content or "")
    budget = re.search(r"预算金额[：:，\s]*([\d,.]+)\s*元?", content or "")
    max_price = re.search(r"最高限价[：:，\s]*([\d,.]+)", content or "")
    return bid_time.group(1) if bid_time else "", budget.group(1) if budget else "", max_price.group(1) if max_price else ""


def fetch_all_procurement_data(start_date, end_date):
    all_records = []
    page = 0
    page_size = 10000
    max_pages = 1
    total_count = None

    while page < max_pages:
        print(f"正在获取第 {page + 1} 页数据...")
        payload = generate_payload(start_date, end_date, page, page_size)
        try:
            response = requests.post(URL, json=payload, headers=HEADERS, timeout=30)
            response.raise_for_status()
            data = response.json()

            if total_count is None:
                total_count = data.get("result", {}).get("totalcount", 0)
                print(f"总记录数: {total_count}")

            records = data.get("result", {}).get("records", [])
            if not records:
                print("没有更多数据了")
                break

            all_records.extend(records)
            print(f"本页获取 {len(records)} 条，累计 {len(all_records)} 条")

            if len(all_records) >= total_count:
                break

            page += 1
            time.sleep(1)
        except Exception as e:
            print(f"第 {page + 1} 页获取失败: {e}")
            break

    return all_records


def build_detail_url(record):
    link = record.get("linkurl")
    if link and link.startswith("/"):
        return "https://www.jxsggzy.cn" + link
    elif record.get("infoid"):
        return f"https://www.jxsggzy.cn/web/jyxx/002006/002006001/{record['infoid']}.html"
    return ""


def set_column_widths_and_style(writer, widths):
    worksheet = writer.sheets["Sheet1"]
    for i, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(i)].width = width

    fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = Font(bold=True)
    worksheet.auto_filter.ref = worksheet.dimensions


def process_and_save_data(records):
    if not records:
        print("未获取到有效数据")
        return False

    filtered = [r for r in records if any(kw in (r.get("title") or r.get("titlenew") or "") for kw in KEYWORDS)]

    df = pd.DataFrame([{
        "发布时间": (r.get("webdate") or "")[:10],
        "辖区": r.get("districtName", r.get("xiaquname", "")),
        "标题": r.get("title") or r.get("titlenew", "无标题"),
        "预算金额": extract_detail_fields(r.get("content", ""))[1],
        "最高限价": extract_detail_fields(r.get("content", ""))[2],
        "投标截止时间": extract_detail_fields(r.get("content", ""))[0],
        "信息类型": r.get("categoryname", "政府采购"),
        "开标类型": r.get("kaibiaotype", ""),
        "网页链接": build_detail_url(r),
        "简要内容": r.get("content", "").lstrip()[:400]
    } for r in filtered])

    filename = f"江西政府采购_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)
        set_column_widths_and_style(writer, [11, 8, 90, 15, 15, 20, 20, 10, 20, 100])

    print(f"\n✅ 成功保存 {len(df)} 条数据到 {filename}")
    print(df.head())
    return True


if __name__ == "__main__":
    print("=== 江西省政府采购数据采集 ===")
    print("Start:", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

    # 设置日期范围：近30天
    end_date = datetime.today().strftime("%Y-%m-%d")
    start_date = (datetime.today() - timedelta(days=30)).strftime("%Y-%m-%d")

    data = fetch_all_procurement_data(start_date, end_date)
    process_and_save_data(data)

    print("End:", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))